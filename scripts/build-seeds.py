#!/usr/bin/env python3
"""Rebuild the bundled JSON seeds from the source xlsx and HTML.

Usage:
    pip3 install openpyxl
    python3 scripts/build-seeds.py <path/to/workout_plan.xlsx> <path/to/mobility-protocol.html>

Writes src/data/workouts.json and src/data/mobility.json.
"""
from __future__ import annotations
import json
import os
import re
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
OUT_DIR = os.path.join(ROOT, 'src', 'data')


def build_workouts(xlsx_path: str) -> dict:
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    DAY_MAP = {
        'Mon – Lower A': ('mon-lower-a', 'Lower A', 'Mon'),
        'Tue – Upper Push': ('tue-push', 'Upper Push', 'Tue'),
        'Wed – Upper Pull': ('wed-pull', 'Upper Pull', 'Wed'),
        'Thu – Lower B': ('thu-lower-b', 'Lower B', 'Thu'),
        'Fri – Upper Full': ('fri-upper-full', 'Upper Full', 'Fri'),
        'Sat – Optional': ('sat-optional', 'Saturday', 'Sat'),
    }

    def slugify(s: str) -> str:
        return re.sub(r'[^a-z0-9]+', '-', s.lower()).strip('-')

    days = []
    for sheet_name, (day_id, display, abbr) in DAY_MAP.items():
        ws = wb[sheet_name]
        rows = [
            list(r)
            for r in ws.iter_rows(values_only=True)
            if any(c is not None and str(c).strip() != '' for c in r)
        ]
        exercises = []
        section = None
        for row in rows:
            first = str(row[0]).strip() if row[0] is not None else ''
            if 'WARMUP' in first:
                section = 'warmup'
                continue
            if 'WORKING SETS' in first or 'LIFTING' in first:
                section = 'working'
                continue
            if 'CORE FINISHER' in first:
                section = 'core'
                continue
            if 'CARDIO' in first:
                section = 'cardio'
                continue
            if first == '#' or first.startswith('Week columns') or first == sheet_name or 'Saturday' in first:
                continue
            if section is None:
                continue
            idx = str(row[0]) if row[0] is not None else ''
            name = str(row[1]).strip() if row[1] is not None else ''
            if not name:
                continue
            if section in ('warmup', 'cardio'):
                ex = {
                    'id': f"{day_id}-{slugify(idx + '-' + name)[:60]}",
                    'dayId': day_id,
                    'section': section,
                    'orderIndex': len(exercises),
                    'name': name,
                    'prescription': str(row[2]).strip() if row[2] is not None else None,
                    'notes': str(row[4]).strip() if len(row) > 4 and row[4] is not None else None,
                }
            else:
                starting = None
                if len(row) > 5 and row[5] is not None:
                    try:
                        starting = float(row[5])
                    except Exception:
                        pass
                ex = {
                    'id': f"{day_id}-{slugify(idx + '-' + name)[:60]}",
                    'dayId': day_id,
                    'section': section,
                    'orderIndex': len(exercises),
                    'name': name,
                    'setsReps': str(row[2]).strip() if row[2] is not None else None,
                    'rest': str(row[3]).strip() if row[3] is not None else None,
                    'notes': str(row[4]).strip() if len(row) > 4 and row[4] is not None else None,
                    'startingWeight': starting,
                }
            exercises.append(ex)
        days.append({'id': day_id, 'display': display, 'abbr': abbr, 'exercises': exercises})

    ref_ws = wb['Reference']
    ref_rows = [list(r) for r in ref_ws.iter_rows(values_only=True)]
    muscle_groups, warmup_rules, execution_reminders = [], [], []
    mode = None
    for row in ref_rows:
        first = str(row[0]).strip() if row[0] is not None else ''
        if 'WEEKLY VOLUME' in first:
            mode = 'vol'; continue
        if 'WARMUP RULES' in first:
            mode = 'warm'; continue
        if 'EXECUTION REMINDERS' in first:
            mode = 'exec'; continue
        if not first or first == 'Muscle Group':
            continue
        if mode == 'vol' and row[1] is not None:
            muscle_groups.append({
                'name': first,
                'setsPerWeek': str(row[1]).strip(),
                'frequency': str(row[2]).strip() if row[2] else '',
            })
        elif mode == 'warm' and row[1] is not None:
            warmup_rules.append({'rule': first, 'detail': str(row[1]).strip()})
        elif mode == 'exec' and row[1] is not None:
            execution_reminders.append({'rule': first, 'detail': str(row[1]).strip()})

    return {
        'days': days,
        'muscleGroups': muscle_groups,
        'warmupRules': warmup_rules,
        'executionReminders': execution_reminders,
    }


def build_mobility(html_path: str) -> dict:
    with open(html_path) as f:
        html = f.read()

    def strip_inner(s: str) -> str:
        return re.sub(r'<(?!/?em\b|/?strong\b)[^>]+>', '', s).strip()

    week_grid = []
    for m in re.finditer(
        r'<div class="day-cell"[^>]*data-day="(\w+)">\s*<div class="abbr">(\w+)</div>\s*<div class="intensity (int-\w+)"></div>\s*<div class="focus">([^<]+)</div>',
        html,
    ):
        day, abbr, intensity_cls, focus = m.groups()
        intensity = {'int-gentle': 'gentle', 'int-mod': 'moderate', 'int-hard': 'hard'}[intensity_cls]
        week_grid.append({'day': day, 'abbr': abbr, 'intensity': intensity, 'focus': focus.strip()})

    sessions = []
    session_re = re.compile(
        r'<details data-day="(\w+)">\s*<summary>.*?<span class="session-title">([^<]+)</span>\s*<span class="session-meta">([^<]+)</span>\s*</summary>\s*<div class="details-body">\s*<div class="session-note">([^<]+)</div>\s*<ul class="exercise-list">(.*?)</ul>',
        re.DOTALL,
    )
    for m in session_re.finditer(html):
        day, title, meta, note, list_html = m.groups()
        exercises = []
        for em in re.finditer(
            r'<li><span class="ex-name">(.*?)</span><span class="ex-dose">([^<]+)</span></li>',
            list_html,
            re.DOTALL,
        ):
            name_html, dose = em.groups()
            href_m = re.search(r'href="#([^"]+)"', name_html)
            ref = href_m.group(1) if href_m else None
            name = re.sub(r'<[^>]+>', '', name_html).strip()
            exercises.append({'name': name, 'dose': dose.strip(), 'ref': ref})
        sessions.append(
            {
                'day': day,
                'title': title.strip().replace('&amp;', '&'),
                'meta': meta.strip(),
                'note': note.strip(),
                'exercises': exercises,
            }
        )

    library = []
    lib_re = re.compile(
        r'<details class="exercise-detail library-item" id="([^"]+)">\s*<summary>.*?<span class="session-title">([^<]+)</span>.*?</summary>\s*<div class="details-body">(.*?)</details>',
        re.DOTALL,
    )
    for m in lib_re.finditer(html):
        lib_id, title, body = m.groups()
        mm = re.search(r'<span><strong>Target</strong>([^<]+)</span>', body)
        target = mm.group(1).strip() if mm else None
        mm = re.search(r'<span><strong>Dose</strong>([^<]+)</span>', body)
        dose = mm.group(1).strip() if mm else None
        sec_list = []
        for sm in re.finditer(
            r'<div class="ex-section"><div class="ex-section-label">([^<]+)</div><p>(.*?)</p></div>',
            body,
            re.DOTALL,
        ):
            sec_list.append({'label': sm.group(1).strip(), 'content': strip_inner(sm.group(2))})
        yt = re.search(r'href="(https://www\.youtube\.com/[^"]+)"', body)
        library.append(
            {
                'id': lib_id,
                'title': title.strip(),
                'target': target,
                'dose': dose,
                'sections': sec_list,
                'youtube': yt.group(1) if yt else None,
            }
        )

    rules = []
    for m in re.finditer(
        r'<div class="rule">\s*<div class="rule-header"><span class="rule-num">(\d+)</span><span class="rule-title">([^<]+)</span></div>\s*<p>(.*?)</p>\s*</div>',
        html,
        re.DOTALL,
    ):
        num, title, text = m.groups()
        rules.append(
            {'num': num, 'title': title.strip().replace('&amp;', '&'), 'text': strip_inner(text)}
        )

    benchmarks = []
    for m in re.finditer(
        r'<div class="bench">\s*<div class="bench-metric">([^<]+)</div>\s*<div class="bench-desc">([^<]+)</div>\s*</div>',
        html,
    ):
        benchmarks.append({'metric': m.group(1).strip(), 'desc': m.group(2).strip()})

    am = re.search(r'<div class="context-note">(.*?)</div>\s*<!--\s*─', html, re.DOTALL)
    ankle = None
    if am:
        body = am.group(1)
        title = re.search(r'<h3>([^<]+)</h3>', body).group(1).strip()
        paras = [
            strip_inner(p) for p in re.findall(r'<p>(.*?)</p>', body, re.DOTALL)
        ]
        items = [
            strip_inner(li) for li in re.findall(r'<li>(.*?)</li>', body, re.DOTALL)
        ]
        ankle = {'title': title, 'paragraphs': paras, 'items': items}

    return {
        'weeklyGrid': week_grid,
        'sessions': sessions,
        'library': library,
        'progressionRules': rules,
        'benchmarks': benchmarks,
        'ankleNote': ankle,
    }


def main():
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(1)
    xlsx, html = sys.argv[1], sys.argv[2]
    with open(os.path.join(OUT_DIR, 'workouts.json'), 'w') as f:
        json.dump(build_workouts(xlsx), f, indent=2)
    with open(os.path.join(OUT_DIR, 'mobility.json'), 'w') as f:
        json.dump(build_mobility(html), f, indent=2)
    print('✓ seeds rebuilt')


if __name__ == '__main__':
    main()
